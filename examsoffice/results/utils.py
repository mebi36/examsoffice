from itertools import chain
import os
from typing import Any, Dict, List

from django.db.models.query import QuerySet
from openpyxl import Workbook, load_workbook
from openpyxl.utils import cell, get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from results.models import Lecturer, Student, Result


NORMAL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TIMES_NEW_ROM_STYLE = Font("Times New Roman", 11)
CENTER_ALIGN = Alignment(horizontal="center")
VERTICAL_ALIGN_TOP = Alignment(vertical="top")
TOP_BOT_TEXT_DIRECTION = Alignment(text_rotation=255, vertical="top")


def _list_to_cell(list_var: List[str]) -> str:
    final_str = ""
    if isinstance(list_var, list) and len(list_var) > 0:
        for index, content in enumerate(list_var, 1):
            final_str += content
            if index != len(list_var):
                final_str += ", "
    else:
        final_str = "No outstanding course for semester"
    return final_str


def _merge_row_wise(
    worksheet: Worksheet, row: int, col_start: int, col_end: int
) -> None:
    cell_start = get_column_letter(col_start) + str(row)
    cell_stop = get_column_letter(col_end) + str(row)
    _range = f"{cell_start}:{cell_stop}"
    worksheet.merge_cells(_range)


def _merge_col_wise(
    worksheet: Worksheet, col: int, row_start: int, row_end: int
) -> None:
    cell_start = get_column_letter(col) + str(row_start)
    cell_stop = get_column_letter(col) + str(row_end)
    _range = f"{cell_start}:{cell_stop}"
    worksheet.merge_cells(_range)


def failed_courses_breakdown(df: DataFrame) -> Dict[str, Any]:
    # collating and writing outstanding/failed courses
    failed_courses = df.loc[df["grade"] == "F"]
    failed_courses = failed_courses["course_code"].tolist()
    failed_courses = list(set(failed_courses))
    failed_courses_copy = failed_courses.copy()
    outstanding_credit_load = 0
    outstanding_cred_1st = 0
    outstanding_cred_2nd = 0
    failed_courses_first = []
    failed_courses_second = []

    for course in failed_courses_copy:
        if df.loc[(df["grade"] != "F") & (df["course_code"] == course)].empty:
            if (
                "First"
                in df.query("course_code == @course")["semester"].iloc[0]
            ):
                outstanding_cred_1st = (
                    outstanding_cred_1st
                    + df.query("course_code == @course")["credit_load"].iloc[0]
                )
                failed_courses_first.append(course)
            else:
                outstanding_cred_2nd = (
                    outstanding_cred_2nd
                    + df.query("course_code == @course")["credit_load"].iloc[0]
                )
                failed_courses_second.append(course)
            outstanding_credit_load = (
                outstanding_credit_load
                + df.query("course_code == @course")["credit_load"].iloc[0]
            )
        else:
            failed_courses.remove(course)

    failed_courses_first.sort(key=lambda x: x[-3:])
    failed_courses_second.sort(key=lambda x: x[-3:])
    failed_courses.sort(key=lambda x: x[-3:])

    return {
        "failed_courses_first": failed_courses_first,
        "failed_courses_second": failed_courses_second,
        "outstanding_cred_1st": outstanding_cred_1st,
        "outstanding_cred_2nd": outstanding_cred_2nd,
        "outstanding_credit_load": outstanding_credit_load,
        "failed_courses": failed_courses,
    }


def student_transcript(transcript_data: Dict[str, Any]) -> Workbook:
    file_path: List[str] = [
        "static",
        "excel_templates",
        "transcript_template.xlsx",
    ]
    wb: Workbook = load_workbook(os.path.join(*file_path))
    ws: Worksheet = wb.active

    # writing the biodata block
    bio_info: List[str] = [
        "ACADEMIC TRANSCRIPT OF",
        "REGISTRATION NUMBER: ",
        "LEVEL OF STUDY: ",
    ]
    row = 9
    col = 5
    # writing result block
    student_bio: List[str] = transcript_data["student_bio"]
    result_columns = [1, 4, 11, 44, 51, 55, 62]
    for idx, el in enumerate(bio_info):
        _ = ws.cell(row=row + idx, column=col, value=el)
        _.font = Font(bold=True)
        _merge_row_wise(ws, row=(row + idx), col_start=col, col_end=(col + 18))

        _ = ws.cell(row=(row + idx), column=(col + 19), value=student_bio[idx])
        _.font = Font(bold=True)
        _.border = Border(bottom=Side(style="thin"))
        _.alignment = Alignment(horizontal="center")
        _merge_row_wise(
            ws, row=row + idx, col_start=(col + 19), col_end=(col + 50)
        )

    row = 13
    col = 1
    result_dict: Dict[str, Dict[str, DataFrame]] = transcript_data[
        "transcript_body"
    ]

    credit_sum = 0
    weight_sum = 0
    for session, results in result_dict.items():
        semesters = [x for x in results.keys()]

        credit_sem_sum = 0
        weight_sem_sum = 0
        for semester in semesters:
            df = results[semester]
            if len(df.index) == 0:
                continue
            res_headings = [
                "SN",
                "Course Code",
                "Course Title",
                "Unit Load",
                "Grade",
                "Grade Point",
            ]

            res_df = df[
                [
                    "course_id__course_code",
                    "course_id__course_title",
                    "course_id__credit_load",
                    "letter_grade",
                    "weight",
                ]
            ]
            semester_title = f"{df['semester_id__desc'].iloc[0]} ({df['course_id__course_level'].max()}/5)"
            credit_sem_sum = df["course_id__credit_load"].sum()
            weight_sem_sum = df["weight"].sum()
            sem_gpa = round(weight_sem_sum / credit_sem_sum, 2)

            credit_sum += credit_sem_sum
            weight_sum += weight_sem_sum

            # preparing result for writing to excel
            df_rows = dataframe_to_rows(res_df, index=False, header=False)

            # printing semester title
            _ = ws.cell(row=row, column=1, value=semester_title)
            _.font = Font(bold=True)
            _.alignment = Alignment(horizontal="center")
            _merge_row_wise(
                ws, row=row, col_start=1, col_end=result_columns[-1]
            )
            row += 1

            for c_idx, header in enumerate(res_headings):
                _ = ws.cell(row=row, column=result_columns[c_idx], value=header)
                _.font = Font(bold=True)
                _.border = NORMAL_BORDER
                _.alignment = Alignment(horizontal="center")
                _merge_row_wise(
                    ws,
                    row=row,
                    col_start=result_columns[c_idx],
                    col_end=(result_columns[c_idx + 1] - 1),
                )
            row += 1

            count = 1
            for r_idx, df_row in enumerate(df_rows, row):
                _ = ws.cell(row=r_idx, column=result_columns[0], value=count)
                _.border = NORMAL_BORDER
                _merge_row_wise(
                    ws,
                    r_idx,
                    col_start=result_columns[0],
                    col_end=(result_columns[1] - 1),
                )
                count += 1
                for c_idx, df_value in enumerate(df_row, 1):
                    _ = ws.cell(
                        row=r_idx,
                        column=result_columns[c_idx],
                        value=str(df_value).upper(),
                    )
                    _.alignment = Alignment(horizontal="center", wrap_text=True)
                    _.border = NORMAL_BORDER

                    _merge_row_wise(
                        ws,
                        r_idx,
                        col_start=result_columns[c_idx],
                        col_end=(result_columns[(c_idx + 1)] - 1),
                    )

            for index, i in enumerate(result_columns):
                _ = ws.cell(row=(r_idx + 1), column=i)
                _.border = NORMAL_BORDER
                if index == 6:
                    _merge_row_wise(
                        ws, row=(r_idx + 1), col_start=i, col_end=(i + 4)
                    )
                else:
                    _merge_row_wise(
                        ws,
                        row=(r_idx + 1),
                        col_start=i,
                        col_end=(result_columns[index + 1] - 1),
                    )
            _ = ws.cell(
                row=(r_idx + 1), column=result_columns[3], value=credit_sem_sum
            )
            _.font = Font(bold=True)
            _.alignment = Alignment(horizontal="center")

            _ = ws.cell(
                row=(r_idx + 1), column=result_columns[5], value=weight_sem_sum
            )
            _.font = Font(bold=True)
            _.alignment = Alignment(horizontal="center")

            _ = ws.cell(
                row=(r_idx + 1), column=result_columns[6], value=sem_gpa
            )
            _.alignment = Alignment(horizontal="center")
            _.font = Font(bold=True)

            row = r_idx + 4
        _ = ws.cell(row=(r_idx + 2), column=result_columns[5], value="CGPA")
        _.font = Font(bold=True)
        _.alignment = Alignment(horizontal="center")

        _ = ws.cell(
            row=(r_idx + 2),
            column=result_columns[6],
            value=round(weight_sum / credit_sum, 2),
        )
        _.font = Font(bold=True)
        _.border = Border(bottom=Side(style="thick"))
        _merge_row_wise(
            ws,
            row=(r_idx + 2),
            col_start=result_columns[6],
            col_end=(result_columns[6] + 4),
        )
        ws.row_breaks.append(Break(id=r_idx + 2))

    hod_name = Lecturer.objects.get(head_of_dept=True).full_name
    ws.HeaderFooter.differentFirst = False
    ws.oddFooter.left.text = "&BHead of Dept.: " + "&U" + hod_name + "  &U&B\n"
    ws.oddFooter.center.text = "&BSign:__________________&B\n"
    ws.oddFooter.right.text = (
        "&BDate:__________________&B\nPage &[Page] of &[Pages]"
    )
    return wb


def class_result_spreadsheet(
    result_qs: QuerySet, class_list: List[str], expected_yr_of_grad: str
) -> Workbook:
    wb = Workbook()
    ws = wb.active

    def _format_summary_block(row: int, col: int) -> None:
        for i in range(row, (row + 4)):
            _ = ws.cell(row=i, column=(col + 20))
            _.fill = PatternFill(fgColor="50E2F2", fill_type="solid")
            _.border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick"),
            )

    def _failed_crses_block(
        course_list: List[str], total_cred: int, first_sem: bool = False
    ) -> None:

        head_fill = "First" if first_sem else "Second"
        f1_head = f"""Courses with no passing grade ({head_fill})"""
        f1_head_cell = ws.cell(row=row - 1, column=(col), value=f1_head)
        _merge_row_wise(ws, row=(row - 1), col_start=col, col_end=(col + 19))
        f1_head_cell.font = Font(bold=True)
        f1_head_cell.alignment = CENTER_ALIGN
        f1_head_cell.fill = PatternFill(fgColor="F7C5C1", fill_type="solid")

        i = 0
        j = 0
        for el in course_list:
            _ = ws.cell(row=(row + i), column=(col + j), value=el)
            _.alignment = Alignment(vertical="top", textRotation=180)
            if i == 0:
                i = 1
            elif i == 1:
                _merge_col_wise(
                    ws, col=(col + j), row_start=(row + i), row_end=(row + 3)
                )

                i = 0
                j = j + 1
        for row_el in range(row, (row + 4)):
            for col_el in range(col, (col + 20)):
                ws.cell(row=row_el, column=col_el).border = NORMAL_BORDER
                ws.column_dimensions[get_column_letter(col_el)].width = 3

        block_title = (
            "Cred Load of CO's (1st Sem)"
            if first_sem
            else "Total Credit Load of CO's"
        )
        _ = ws.cell(row=row, column=(col + 20), value=block_title)
        _.alignment = Alignment(wrap_text=True)
        ws.cell(row=(row + 1), column=(col + 20), value=total_cred)
        _format_summary_block(row, col)

    ws.title = f"Class of {expected_yr_of_grad} Results"
    headers = ["SN", "Name", "RN", "E.M"]
    ws.append(headers)
    ws.freeze_panes = "E1"
    serial_number = 1
    row = 2
    class_list.sort(key=lambda x: x[1])

    for student in class_list:
        student_qs = result_qs.filter(student_reg_no=student[0])
        # confirm that the student's results are available
        if len(student_qs) == 0:
            continue

        else:
            df = DataFrame.from_records(student_qs)
            df = df.rename(
                columns={
                    0: "course_title",
                    1: "course_code",
                    2: "credit_load",
                    3: "course_level",
                    4: "semester",
                    5: "grade",
                }
            )
            df = Student.get_weight_col(df)
            reg_no = student[0]
            name = student[1]
            mode_of_admission = student[2] or "N/A"
            semesters = df["semester"].unique().tolist()
            semesters = sorted(semesters)

            # writing and formatting student biodata to the excel sheet
            sn_cell = ws.cell(column=1, row=row, value=serial_number)
            sn_cell.alignment = TOP_BOT_TEXT_DIRECTION
            sn_cell.border = NORMAL_BORDER
            _merge_col_wise(ws, col=1, row_start=row, row_end=row + 3)

            name_cell = ws.cell(column=2, row=row, value=name.upper())
            name_cell.alignment = Alignment(wrapText=True, vertical="top")
            name_cell.border = NORMAL_BORDER
            _merge_col_wise(ws, col=2, row_start=row, row_end=row + 3)
            ws.column_dimensions["B"].width = 20

            reg_no_cell = ws.cell(column=3, row=row, value=reg_no)
            reg_no_cell.alignment = Alignment(textRotation=180, vertical="top")
            reg_no_cell.border = NORMAL_BORDER
            _merge_col_wise(ws, col=3, row_start=row, row_end=row + 3)

            adm_mode_cell = ws.cell(
                column=4, row=row, value=mode_of_admission.upper()
            )
            adm_mode_cell.alignment = Alignment(
                textRotation=180, vertical="top"
            )
            adm_mode_cell.border = NORMAL_BORDER
            _merge_col_wise(ws, col=4, row_start=row, row_end=row + 3)

            for el in ["A", "C", "D"]:
                ws.column_dimensions[el].width = 3

            # writing results to worksheet
            cred_sum = 0
            weight_sum = 0
            col = 5

            semester = semesters.extend(["", ""])
            for semester in semesters:
                semester_records = df.query("semester == @semester")
                semester_records = semester_records.sort_values(
                    "course_level", ascending=False
                )
                semester_records = semester_records[
                    ["course_code", "credit_load", "grade", "weight"]
                ].copy()
                semester_records = semester_records.transpose()

                # calculating the semester CGPA& and writing to the
                # last column
                credit_sem_sum = semester_records.sum(axis=1)[1]
                weight_sem_sum = semester_records.sum(axis=1)[3]
                weight_sum = weight_sum + weight_sem_sum
                cred_sum = cred_sum + credit_sem_sum

                # Running CGPA
                col_tally = col
                cgpa_num = f"{get_column_letter(col+20)}{row+3}"
                cgpa_denom = f"{get_column_letter(col+20)}{row+1}"
                while (col_tally - 20) > 0:
                    cgpa_num += f"+{get_column_letter(col_tally-1)}{row+3}"
                    cgpa_denom += f"+{get_column_letter(col_tally-1)}{row+1}"
                    col_tally -= 20
                    if col_tally < 10:
                        break
                ws.cell(
                    row=row,
                    column=(col + 20),
                    value=(
                        f'=IF(ISBLANK({get_column_letter(col)}{row+2}),"",'
                        f'ROUND((({cgpa_num})/({cgpa_denom})),3))'
                    )
                )
                # Total Semester Grade Point cell
                sem_grade_pt_formula = (
                    f"=SUM({get_column_letter(col)}{row+1}"
                    f":{get_column_letter(col+19)}{row+1})"
                )
                ws.cell(
                    row=(row + 1),
                    column=(col + 20),
                    value=sem_grade_pt_formula
                )
                # Semester GPA cell
                sem_cgpa_formula = (
                    f'=IF(ISBLANK({get_column_letter(col)}{row+2}),"",ROUND('
                    f'SUM({get_column_letter(col)}{row+3}:'
                    f'{get_column_letter(col+19)}{row+3})/'
                    f'SUM({get_column_letter(col)}{row+1}:'
                    f'{get_column_letter(col+19)}{row+1}), 3))'
                )
                ws.cell(
                    row=(row + 2),
                    column=(col + 20),
                    value=sem_cgpa_formula
                )
                # Total semester credit load cell
                sem_cred_formula = (
                    f"=SUM({get_column_letter(col)}{row+3}:"
                    f"{get_column_letter(col+19)}{row+3})"
                )
                ws.cell(
                    row=(row + 3), column=(col + 20), value=sem_cred_formula
                )
                _format_summary_block(row, col)

                # prepare result dataframe for writing to excel
                df_rows = dataframe_to_rows(
                    semester_records, index=False, header=False
                )

                # working on the semester header
                ws.cell(column=col, row=(row - 1), value=semester.upper())
                header_start = get_column_letter(col) + str(row - 1)
                _merge_row_wise(
                    ws, row=(row - 1), col_start=col, col_end=(col + 19)
                )
                ws[header_start].alignment = CENTER_ALIGN
                ws[header_start].fill = PatternFill(
                    fgColor="50E2F2", fill_type="solid"
                )
                ws[header_start].font = Font(bold=True)

                for r_idx, df_row in enumerate(df_rows, row):
                    for all_col in range(col, (col + 20)):
                        _ = ws.cell(column=all_col, row=r_idx)
                        _.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )
                        ws.column_dimensions[
                            get_column_letter(all_col)
                        ].width = 3
                    for c_idx, df_value in enumerate(df_row, col):
                        _ = ws.cell(column=c_idx, row=r_idx, value=df_value)
                        _.alignment = Alignment(horizontal="center")
                        if r_idx == row:
                            _.alignment = Alignment(
                                text_rotation=180, vertical="top"
                            )
                # Grade Point cells
                for c in range(20):
                    grd = get_column_letter(col+c) + str(row+2)  # letter grade cell
                    un_ld = get_column_letter(col+c) + str(row+1)  # Unit load cell
                    ws.cell(
                        row=row+3,
                        column=col+c,
                        value=(
                            f'=IF(ISBLANK({grd}),"",IF(({grd}="A"),"5",'
                            f'IF(({grd}="B"),"4",IF(({grd}="C"),"3",'
                            f'IF(({grd}="D"),"2",IF(({grd}="E"),"1",'
                            f'IF(({grd}="F"),"0"," "))))))*{un_ld})'
                        )
                    )

                col += 21

            # collating and writing outstanding/failed courses
            course_brk_dwn = failed_courses_breakdown(df)
            failed_courses_first = course_brk_dwn["failed_courses_first"]
            outstanding_cred_1st = course_brk_dwn["outstanding_cred_1st"]
            outstanding_cred = course_brk_dwn["outstanding_credit_load"]
            failed_courses_second = course_brk_dwn["failed_courses_second"]

            _failed_crses_block(
                failed_courses_first, outstanding_cred_1st, first_sem=True
            )
            col += 21

            _failed_crses_block(failed_courses_second, outstanding_cred)

            row += 5
            serial_number += 1

    return wb


def class_failure_spreadsheet(
    result_qs: QuerySet, class_list: List[str], expected_yr_of_grad: str
) -> Worksheet:
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active

    ws.title = f"Class of {expected_yr_of_grad} CO Summary List"
    headers = [
        "SN",
        "Reg Number",
        "Name",
        "Outstanding Courses (1st)",
        "Credit Load(1st)",
        "Outstanding Courses (2nd)",
        "Total Credit Load",
        "CGPA",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    serial_number = 1
    column_dimensions = {1: 4, 2: 12, 3: 20, 4: 28, 5: 10, 6: 28, 7: 11, 8: 9}

    for key, value in column_dimensions.items():
        ws.column_dimensions[get_column_letter(key)].width = value
        cell = ws.cell(row=1, column=key)
        cell.border = NORMAL_BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="top", wrap_text=True
        )
        cell.font = Font(bold=True)

    row = 2
    for student in class_list:
        student_qs = result_qs.filter(student_reg_no=student[0]).order_by(
            "semester"
        )
        # confirm that the student's results are available
        if len(student_qs) == 0:
            continue

        else:
            # convert student's result qs to dataframe
            df = DataFrame.from_records(student_qs)
            df = df.rename(
                columns={
                    0: "course_title",
                    1: "course_code",
                    2: "credit_load",
                    3: "course_level",
                    4: "semester",
                    5: "grade",
                }
            )
            df = Student.get_weight_col(df)
            reg_no = student[0]
            name = student[1]
            # df.sort_values('course_level', ascending=False, in_place=True)
            # calculate the cgpa of the student
            weight_sum = df["weight"].sum()
            credit_sum = df["credit_load"].sum()
            cgpa = round(weight_sum / credit_sum, 3)

            # get dictionary containing breakdown of failed courses
            course_brk_dwn = failed_courses_breakdown(df)
            failed_courses_first = course_brk_dwn["failed_courses_first"]
            outstanding_cred_1st = course_brk_dwn["outstanding_cred_1st"]
            outstanding_cred = course_brk_dwn["outstanding_credit_load"]
            failed_courses_second = course_brk_dwn["failed_courses_second"]

            entry_alignment = Alignment(
                vertical="top", horizontal="center", wrap_text=True
            )

            ws.row_dimensions[row].height = 90
            ws.cell(row=row, column=1, value=serial_number)
            ws.cell(row=row, column=2, value=reg_no)
            ws.cell(row=row, column=3, value=name)
            ws.cell(
                row=row, column=4, value=_list_to_cell(failed_courses_first)
            )
            ws.cell(row=row, column=5, value=outstanding_cred_1st)
            ws.cell(
                row=row, column=6, value=_list_to_cell(failed_courses_second)
            )
            ws.cell(row=row, column=7, value=outstanding_cred)
            ws.cell(row=row, column=8, value=cgpa)

            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = entry_alignment
                cell.border = NORMAL_BORDER

            row += 1
            serial_number += 1
    ws.oddHeader.center.text = """&BDEPARTMENT OF ELECTRONIC ENGINEERING\nSTUDENT ACADEMIC PERFORMANCE SUMMARY&B"""
    Worksheet.set_printer_settings(ws, paper_size=9, orientation="landscape")
    return wb


def class_of_degree_spreadsheet(
    result_qs: QuerySet, class_list: List[str], expected_yr_of_grad: str
) -> Worksheet:
    classes_of_degree = {
        4.5: "1st Class",
        3.5: "2nd Class (Upper)",
        2.5: "2nd Class (Lower)",
        1.5: "3rd Class",
        0: "Pass"
    }
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active

    ws.title = f"Class of {expected_yr_of_grad} Potential Class of Degree"
    headers = [
        "SN",
        "Reg Number",
        "Name",
        "Potential Class of Degree"
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    serial_number = 1
    column_dimensions = {1: 4, 2: 18, 3: 26, 4: 12}

    for key, value in column_dimensions.items():
        ws.column_dimensions[get_column_letter(key)].width = value
        cell = ws.cell(row=1, column=key)
        cell.border = NORMAL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.font = Font(bold=True)
    
    row = 2
    # class_list = class_list.order_by("last_name")
    class_list = sorted(class_list, key=lambda x: x[1])
    for student in class_list:
        student_qs = result_qs.filter(student_reg_no=student[0]).order_by(
            "semester"
        )
        # confirm that the student's results are available
        if len(student_qs) == 0:
            continue

        else:
            # convert student's result qs to dataframe
            df = DataFrame.from_records(student_qs)
            df = df.rename(
                columns={
                    0: "course_title",
                    1: "course_code",
                    2: "credit_load",
                    3: "course_level",
                    4: "semester",
                    5: "grade",
                }
            )
            df = Student.get_weight_col(df)
            reg_no = student[0]
            name = student[1]
            weight_sum = df["weight"].sum()
            credit_sum = df["credit_load"].sum()
            cgpa = round(weight_sum / credit_sum, 3)

            for key, val in classes_of_degree.items():
                if cgpa >= key:
                    class_of_deg = val
                    break
        
            # get failed failed courses
            failed_courses = failed_courses_breakdown(df)["failed_courses"]

            if len(failed_courses) != 0:
                continue
            
            entry_alignment = Alignment(vertical="top", horizontal="center")
            ws.row_dimensions[row].height = 90
            ws.cell(row=row, column=1, value=serial_number)
            ws.cell(row=row, column=2, value=reg_no)
            ws.cell(row=row, column=3, value=name)
            ws.cell(row=row, column=4, value=class_of_deg)

            for col_idx in range(1, 5):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = entry_alignment
                cell.border = NORMAL_BORDER
            
            row += 1
            serial_number += 1
    
    ws.oddHeader.center.text = (
        f'''&BDEPARTMENT OF ELECTRONIC ENGINEERING_'''
        f'''POTENTIAL CLASS OF DEGREE ({expected_yr_of_grad})&B'''
    )
    ws.oddFooter.center.text = (
        '''&08&IThis document is only a projection and will only become '''
        '''official when the results of the students presented herewith'''
        ''' have been ratified by the University's examination unit.&I'''
    )
    return wb


def collated_results_spreadsheet(res_df: DataFrame) -> Worksheet:
    """This function will process the result df into a spreadsheet which
    will be returned to the caller"""

    courses = res_df["course_code"].unique().tolist()
    students = res_df.groupby(["name", "reg_no"]).size().reset_index()
    students = students[["name", "reg_no"]].values.tolist()
    collated_result = DataFrame(students, columns=["Name", "RegNo"])
    for course in courses:
        x = ["X"] * len(students)
        j = course
        for student in range(0, len(students)):
            grade = res_df.query(
                "(reg_no == @students[@student][1]) and (course_code == @course)"
            )
            if len(grade) == 1:
                x[student] = grade["grade"].tolist()[0]
        collated_result[j] = x

    # printing to a worksheet
    wb = Workbook()
    ws = wb.active

    row = 1
    col = 1
    ws.freeze_panes = "D2"
    headings = ["SN", "Name", "REG NO"] + courses

    for c_idx in chain(range(1, 2), range(4, (len(headings) + 1))):
        ws.column_dimensions[get_column_letter(c_idx)].width = 4
    ws.column_dimensions[get_column_letter(2)].width = 38
    ws.column_dimensions[get_column_letter(3)].width = 12
    for idx, heading in enumerate(headings, 1):
        _ = ws.cell(row=1, column=idx, value=heading)
        _.font = Font("Times new Roman", 13, bold=True)
        _.border = NORMAL_BORDER
        _.alignment = Alignment(vertical="top", horizontal="center")
        if idx > 3:
            _.alignment = Alignment(
                vertical="top", text_rotation=180, horizontal="center"
            )
    row += 1

    result_rows = dataframe_to_rows(collated_result, index=False, header=False)
    count = 1
    for row_idx, df_row in enumerate(result_rows, row):
        _ = ws.cell(row=row_idx, column=1, value=count)
        _.border = NORMAL_BORDER
        _.font = TIMES_NEW_ROM_STYLE
        _.alignment = Alignment(horizontal="center", vertical="top")
        count += 1
        for col_idx, df_value in enumerate(df_row, col + 1):
            _ = ws.cell(row=row_idx, column=col_idx, value=df_value.upper())
            _.border = NORMAL_BORDER
            _.alignment = Alignment(
                horizontal="center", vertical="top", wrap_text=True
            )
            _.font = TIMES_NEW_ROM_STYLE

    hod_name = Lecturer.objects.get(head_of_dept=True).full_name

    ws.print_title_rows = "1:1"
    ws.HeaderFooter.differentFirst = False
    ws.oddFooter.left.text = (
        "&BHead of Dept.: "
        + "&U"
        + hod_name
        + "  &U&B\nSign:__________________"
    )
    ws.oddFooter.right.text = (
        "&BExams Officer:__________________&B\nSign:__________________"
    )
    Worksheet.set_printer_settings(ws, paper_size=9, orientation="landscape")
    return wb


def possible_graduands_wb(expected_yr_of_grad: str) -> Worksheet:
    """Returns a worksheet of possible graduands for a graduation year."""
    student_qs: QuerySet = (
        Student.objects.all()
        .filter(expected_yr_of_grad=expected_yr_of_grad)
        .values(
            "student_reg_no",
            "jamb_number",
            "phone_number",
            "last_name",
            "first_name",
            "other_names",
            "sex__sex",
            "marital_status__marital_status",
            "lga_of_origin",
            "state_of_origin",
            "email",
            "current_level_of_study",
            "level_admitted_to",
            "date_of_birth",
        )
    )
    if not student_qs.exists():
        raise ValueError(
            f"No student with expected year of graduation {expected_yr_of_grad}"
        )
    bio_df = DataFrame(student_qs)
    class_list = bio_df["student_reg_no"].to_list()
    result_qs = (
        Result.objects.all()
        .select_related("semester", "course")
        .filter(student_reg_no__in=class_list)
        .values_list(
            "course_id__course_title",
            "course_id__course_code",
            "course_id__credit_load",
            "course_id__course_level",
            "semester_id__desc",
            "letter_grade",
        )
    )

    eligible_students: List[str] = []

    # checking eligibility of every student in the class list
    for student in class_list:
        student_qs = result_qs.filter(student_reg_no=student).order_by(
            "semester"
        )

        if not student_qs.exists():
            continue

        else:
            df = DataFrame.from_records(student_qs)
            df = df.rename(
                columns={
                    0: "course_title",
                    1: "course_code",
                    2: "credit_load",
                    3: "course_level",
                    4: "semester",
                    5: "grade",
                }
            )

            df = Student.get_weight_col(df)
            course_brk_dwn = failed_courses_breakdown(df)
            outstanding_cred_1st = course_brk_dwn["outstanding_cred_1st"]
            outstanding_cred_2nd = course_brk_dwn["outstanding_cred_2nd"]
            reg_courses_load_1st = 19
            reg_course_load_2nd = 16

            if (reg_courses_load_1st + outstanding_cred_1st) <= 24 and (
                reg_course_load_2nd + outstanding_cred_2nd
            ) <= 24:
                eligible_students.append(student)

    grad_df = bio_df[bio_df["student_reg_no"].isin(eligible_students)]

    # cleaning up some columns of the df
    grad_df["date_of_birth"] = grad_df["date_of_birth"].dt.strftime("%d/%m/%Y")
    grad_df["jamb_no_phone_no"] = (
        grad_df["jamb_number"].fillna("")
        + "/+234"
        + grad_df["phone_number"].fillna("--")
    )
    grad_df["full_name"] = (
        grad_df["last_name"].fillna("")
        + " "
        + grad_df["first_name"].fillna("")
        + " "
        + grad_df["other_names"].fillna("")
    )
    grad_df["department"] = "Electronic Engineering"
    grad_df["duration"] = (
        grad_df["current_level_of_study"].fillna(5)
        - grad_df["level_admitted_to"].fillna(1)
    ).astype(int) + 1
    grad_df["signature"] = " "
    grad_df["sex_abbr"] = grad_df.sex__sex.str[:1]
    req_cols = [
        "student_reg_no",
        "jamb_no_phone_no",
        "department",
        "full_name",
        "sex_abbr",
        "marital_status__marital_status",
        "lga_of_origin",
        "state_of_origin",
        "email",
        "duration",
        "date_of_birth",
        "signature",
    ]
    grad_df = grad_df[req_cols]
    grad_df.sort_values("full_name", inplace=True)

    # writing to excel workbook
    wb = Workbook()
    ws = wb.active
    df_rows = dataframe_to_rows(grad_df, index=False, header=False)

    ws_headings = [
        "S/N",
        "Reg. No.",
        "Jamb No./Phone No.",
        "Department",
        "Full Name",
        "Sex",
        "Marital Status",
        "L.G.A",
        "State of Origin",
        "Email",
        "Duration",
        "Date of Birth",
        "Signature",
    ]
    ws_title = f"LIST OF {int(expected_yr_of_grad)-1}/{expected_yr_of_grad} POSSIBLE GRADUANDS"
    count = 1
    row = 5

    _ = ws.cell(row=row, column=1, value=ws_title)
    _.alignment = CENTER_ALIGN
    _.font = Font(bold=True)
    _merge_row_wise(ws, row, 1, 13)
    row += 2
    ws.freeze_panes = "E1"
    for c_idx, header in enumerate(ws_headings, 1):
        _ = ws.cell(row=row, column=c_idx, value=header)
        _.font = Font(bold=True)
        _.border = NORMAL_BORDER
        _.alignment = Alignment(
            horizontal="center", wrap_text=True, vertical="top"
        )
    row += 1

    for r_idx, df_row in enumerate(df_rows, row):
        _ = ws.cell(row=r_idx, column=1, value=count)
        _.border = NORMAL_BORDER
        _.alignment = Alignment(vertical="top")
        count += 1
        for c_idx, df_value in enumerate(df_row, 2):
            cell_val = (
                str(df_value).upper()
                if (str(df_value).upper() not in ["NONE", "NAT", "NAN"])
                else ""
            )
            _ = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            _.alignment = Alignment(
                horizontal="center", wrap_text=True, vertical="top"
            )
            _.border = NORMAL_BORDER
    ws.print_title_rows = "7:7"
    ws.HeaderFooter.differentFirst = True
    ws.firstHeader.center.text = """&BUNIVERSITY OF NIGERIA, NSUKKA\nOFFICE OF THE REGISTRAR\n(EXAMINATIONS)\n\nFACULTY/SCHOOL: ____&UENGINEERING&U____  DEPARTMENT/COURSE COMBINATION: ____&UELECTRONIC&U____&B"""
    ws.firstHeader.size = 16
    Worksheet.set_printer_settings(ws, paper_size=9, orientation="landscape")
    return wb
